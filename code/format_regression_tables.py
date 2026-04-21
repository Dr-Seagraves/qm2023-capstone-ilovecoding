"""
QM 2023 Capstone — Milestone 3: Academic Regression Table Formatting
Creates publication-ready regression tables with standard academic formatting.

Features:
  - One column per model
  - Variables listed on left
  - Coefficients with standard errors in parentheses
  - Summary statistics at bottom (N, R², F-stat, fixed effects, clustering info)
  - Saved as CSV and Excel for ease of review
"""

import os
import sys
import pandas as pd
import numpy as np
from pathlib import Path

# Add config to path
sys.path.insert(0, str(Path(__file__).parent))
from config_paths import FINAL_DATA_DIR, FIGURES_DIR, TABLES_DIR

print("\n" + "="*80)
print("FORMATTING REGRESSION TABLES FOR ACADEMIC PRESENTATION")
print("="*80)

# ============================================================================
# SECTION 1: LOAD RAW MODEL RESULTS
# ============================================================================

print("\n[1/3] Loading model results...")

# Model A (Fixed Effects)
model_a_path = TABLES_DIR / 'M3_model_A_results.csv'
if model_a_path.exists():
    model_a = pd.read_csv(model_a_path)
    print(f"  ✓ Loaded Model A: {model_a_path.name}")
else:
    print(f"  ✗ WARNING: {model_a_path.name} not found. Run capstone_models.py first.")
    model_a = None

# Model B (DiD)
model_b_path = TABLES_DIR / 'M3_model_B_results.csv'
if model_b_path.exists():
    model_b = pd.read_csv(model_b_path)
    print(f"  ✓ Loaded Model B: {model_b_path.name}")
else:
    print(f"  ✗ WARNING: {model_b_path.name} not found. Run capstone_models.py first.")
    model_b = None

# ============================================================================
# SECTION 2: FORMAT COEFFICIENTS WITH STANDARD ERRORS
# ============================================================================

print("\n[2/3] Formatting table with coefficients & standard errors...")

def format_coefficient_with_se(coef, se, pval):
    """
    Format coefficient with standard error in parentheses.
    Add significance stars based on p-value.
    """
    if pd.isna(coef) or pd.isna(se):
        return ""
    
    # Determine significance stars
    if pval < 0.001:
        stars = "***"
    elif pval < 0.01:
        stars = "**"
    elif pval < 0.05:
        stars = "*"
    elif pval < 0.10:
        stars = "†"  # Marginal significance
    else:
        stars = ""
    
    # Format: coefficient with stars on same line, SE in parentheses below
    coef_str = f"{coef:.4f}{stars}"
    se_str = f"({se:.4f})"
    return coef_str, se_str


# Create formatted table
formatted_data = []

if model_a is not None:
    print("  • Formatting Model A (Two-Way Fixed Effects)...")
    
    for idx, row in model_a.iterrows():
        var_name = row['Variable']
        coef, se, pval = row['Coefficient'], row['Std_Error'], row['p_value']
        
        coef_str, se_str = format_coefficient_with_se(coef, se, pval)
        
        formatted_data.append({
            'Variable': var_name,
            'Model A (FE)\nCoefficient': coef_str,
            'Model A (FE)\nStd. Error': se_str,
            'Model A\np-value': f"{pval:.4f}",
        })

if model_b is not None:
    print("  • Formatting Model B (Difference-in-Differences)...")
    
    # Add Model B data to the same rows (matching variable names where possible)
    for idx_b, row_b in model_b.iterrows():
        var_name_b = row_b['Variable']
        coef_b, se_b, pval_b = row_b['Coefficient'], row_b['Std_Error'], row_b['p_value']
        
        coef_str_b, se_str_b = format_coefficient_with_se(coef_b, se_b, pval_b)
        
        # Find if this variable already exists (for variables in both models)
        found = False
        for item in formatted_data:
            if item['Variable'] == var_name_b:
                item['Model B (DiD)\nCoefficient'] = coef_str_b
                item['Model B (DiD)\nStd. Error'] = se_str_b
                item['Model B\np-value'] = f"{pval_b:.4f}"
                found = True
                break
        
        # If not found, add new row
        if not found:
            formatted_data.append({
                'Variable': var_name_b,
                'Model A (FE)\nCoefficient': '',
                'Model A (FE)\nStd. Error': '',
                'Model A\np-value': '',
                'Model B (DiD)\nCoefficient': coef_str_b,
                'Model B (DiD)\nStd. Error': se_str_b,
                'Model B\np-value': f"{pval_b:.4f}",
            })

# Convert to DataFrame
formatted_table = pd.DataFrame(formatted_data)

# ============================================================================
# SECTION 3: ADD SUMMARY STATISTICS
# ============================================================================

print("  • Adding summary statistics...")

# Summary rows (hardcoded from M3 findings report)
summary_stats = [
    {'Variable': '', 'Model A (FE)\nCoefficient': '', 'Model A (FE)\nStd. Error': '', 'Model A\np-value': '', 'Model B (DiD)\nCoefficient': '', 'Model B (DiD)\nStd. Error': '', 'Model B\np-value': ''},  # Blank row
    {'Variable': 'Observations', 'Model A (FE)\nCoefficient': '33,573', 'Model A (FE)\nStd. Error': '', 'Model A\np-value': '', 'Model B (DiD)\nCoefficient': '33,487', 'Model B (DiD)\nStd. Error': '', 'Model B\np-value': ''},
    {'Variable': 'Unique REITs', 'Model A (FE)\nCoefficient': '273', 'Model A (FE)\nStd. Error': '', 'Model A\np-value': '', 'Model B (DiD)\nCoefficient': '273', 'Model B (DiD)\nStd. Error': '', 'Model B\np-value': ''},
    {'Variable': 'Time Periods (Months)', 'Model A (FE)\nCoefficient': '299', 'Model A (FE)\nStd. Error': '', 'Model A\np-value': '', 'Model B (DiD)\nCoefficient': '299', 'Model B (DiD)\nStd. Error': '', 'Model B\np-value': ''},
    {'Variable': 'R² Within', 'Model A (FE)\nCoefficient': '−0.0006', 'Model A (FE)\nStd. Error': '', 'Model A\np-value': '', 'Model B (DiD)\nCoefficient': '0.0015', 'Model B (DiD)\nStd. Error': '', 'Model B\np-value': ''},
    {'Variable': 'R² Between', 'Model A (FE)\nCoefficient': '0.2271', 'Model A (FE)\nStd. Error': '', 'Model A\np-value': '', 'Model B (DiD)\nCoefficient': '0.0089', 'Model B (DiD)\nStd. Error': '', 'Model B\np-value': ''},
    {'Variable': 'R² Overall', 'Model A (FE)\nCoefficient': '0.0169', 'Model A (FE)\nStd. Error': '', 'Model A\np-value': '', 'Model B (DiD)\nCoefficient': '0.0042', 'Model B (DiD)\nStd. Error': '', 'Model B\np-value': ''},
    {'Variable': 'F-statistic', 'Model A (FE)\nCoefficient': '8.23***', 'Model A (FE)\nStd. Error': '', 'Model A\np-value': '<0.001', 'Model B (DiD)\nCoefficient': '12.57***', 'Model B (DiD)\nStd. Error': '', 'Model B\np-value': '<0.001'},
    {'Variable': 'Entity Fixed Effects', 'Model A (FE)\nCoefficient': 'Yes', 'Model A (FE)\nStd. Error': '', 'Model A\np-value': '', 'Model B (DiD)\nCoefficient': 'No', 'Model B (DiD)\nStd. Error': '', 'Model B\np-value': ''},
    {'Variable': 'Time Fixed Effects', 'Model A (FE)\nCoefficient': 'Yes', 'Model A (FE)\nStd. Error': '', 'Model A\np-value': '', 'Model B (DiD)\nCoefficient': 'No', 'Model B (DiD)\nStd. Error': '', 'Model B\np-value': ''},
    {'Variable': 'Clustered SE (by Entity)', 'Model A (FE)\nCoefficient': 'Yes', 'Model A (FE)\nStd. Error': '', 'Model A\np-value': '', 'Model B (DiD)\nCoefficient': 'Yes', 'Model B (DiD)\nStd. Error': '', 'Model B\np-value': ''},
]

summary_df = pd.DataFrame(summary_stats)
formatted_table_with_summary = pd.concat([formatted_table, summary_df], ignore_index=True)

# ============================================================================
# SECTION 4: SAVE AS CSV AND EXCEL
# ============================================================================

print("\n[3/3] Saving formatted tables...")

# CSV format (for easy review)
csv_path = TABLES_DIR / 'M3_REGRESSION_TABLE_FORMATTED.csv'
formatted_table_with_summary.to_csv(csv_path, index=False)
print(f"  ✓ CSV Table saved: {csv_path.name}")

# Excel format (with better formatting)
try:
    excel_path = TABLES_DIR / 'M3_REGRESSION_TABLE_FORMATTED.xlsx'
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        formatted_table_with_summary.to_excel(writer, sheet_name='Regression Results', index=False)
    
    # Optional: Apply formatting (requires openpyxl)
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True, size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Column widths
        ws.column_dimensions['A'].width = 20
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            ws.column_dimensions[col].width = 18
        
        wb.save(excel_path)
        print(f"  ✓ Excel Table saved: {excel_path.name} (with formatting)")
    except ImportError:
        print(f"  ✓ Excel Table saved: {excel_path.name} (basic format)")
        
except Exception as e:
    print(f"  ⚠️  Could not save Excel file: {e}")

# ============================================================================
# SUMMARY & NOTES
# ============================================================================

print(f"\n" + "="*80)
print("FORMATTING COMPLETE")
print("="*80)

print(f"\nFormatted Tables Saved:")
print(f"  • {csv_path.absolute()}")
print(f"  • {excel_path.absolute() if 'excel_path' in locals() else 'Excel version not created'}")

print(f"\nTable Format:")
print(f"  • One column per model (Model A: Two-Way FE; Model B: DiD)")
print(f"  • Coefficients with significance stars: *** p<0.01, ** p<0.05, * p<0.10, † p<0.10")
print(f"  • Standard errors in parentheses below each coefficient")
print(f"  • Summary statistics at bottom: N, R², F-stat, fixed effects info, clustering info")

print(f"\nSignificance Notation:")
print(f"  *** : p < 0.001 (highly significant)")
print(f"  **  : p < 0.01  (very significant)")
print(f"  *   : p < 0.05  (significant)")
print(f"  †   : p < 0.10  (marginally significant)")
print(f"  (blank) : p ≥ 0.10 (not significant)")

print(f"\nUse these tables in:")
print(f"  • Presentations and slide decks")
print(f"  • Academic papers or reports")
print(f"  • Stakeholder communications")
